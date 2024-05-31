
# Импорт библиотеки openpyxl для чтения и записи файлов формата .xlsx.
import openpyxl


# Строка приветствия пользователя.
greetings = """Привет, дорогой друг!\nС помощью данной программы, ты сможешь узнать ответ на главный вопрос!
А именно: какой же коэффициент рассеянной энергии и жёсткость у прекрасных изделий из металл-резины!
Ну что же, не будем тянуть время и приступим!
Тебе нужно будет ответить на несколько вопросов."""

# Строка ознакомления пользователя с функционаллом программы.
familiarization = """Вначале программа попросит вас передать ссылку на файл с расширением .xlsx, 
в котором хранятся данные для расчетов, либо полный путь из корневой папки, либо прямой путь, 
если файл находиться в одной директории или дочерней дириектории относительно программы.
После чего вам предоставляется выбор из двух вариантов:

1. Расчитать и перенести данные коэффициента рассеивания и жёсткости в новый файл.
Здесь так же будут выбор, а именно, произветси расчёты для всех листов в файле, либо только части из них.
Программа попросит вас ввести целое число. Нумерация начинается с первого листа, 
так же если вы укажите большее количество, чем то которое находится в файле, 
программма просто произведёт расчёты для всех листов из файла. В случае если вы укажите нулевое, 
либо отрицательное значение, программа не поймёт вас и попросит повторить ввод.
По мере произвдения расчётов для каждого листа, вы будете видеть индикацию расчётов, а также, 
оценку программы касательно производимых ею расчётов. Зелёная галочка - корректно, 
красный восклицательный знак - возможны отклонения, чёрный крестик - расчет для данного листа не валиден 
(здесь вам может помочь второй способ взаимодействия с программой).
После всего вышеупомянутого, программа предложит дать новому файлу имя (расширене xlsx проставится автоматически) 
и предложит продолжить, либо выйти.

2. Посмотеть всю возможную информацию по интересующему вас листу из файла.
А именно: начало и конец полки, точки экстремумов и антиподов, расчёт коэффициента рассеивания и жёстости.
Надеюсь, данное введение было понятно и вам понравится пользоваться данным творением!)

P.S. За более подробной информацие, пишите по адресу: 'https://t.me/EvgenBojarov'\n"""


def read_and_write_xlsx(link: str, num_sheet: int) -> (list[float], list[float]):
    """
    Функция для четения файла формата xlsx и запись его данных по отдельным именнованым контейнерам типа list.
    :param link: Ссылка на файл.
    :param num_sheet: Номер листа для чтения.
    :return: Два именнованных контейнера с данными об удлинении и нагрузки испытуемого образца.
    """
    elongation: list = []  # Удлинение (см)
    burden: list = []  # Нагрузка (N)
    book = openpyxl.open(link, read_only=True)  # Объект принимающий ссылку с целью чтения.
    sheet = book.worksheets[num_sheet]  # Объект для уточнения листа для чтения.
    # Цикл, на запись данных по контейнерам.
    for e, b in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        elongation.append(abs(e.value))
        burden.append(abs(b.value))
    return elongation, burden


def search_shelf(burden: list[float]) -> tuple[int, int]:
    """
    Функция для нахождения начала и конца полки.
    :param burden: Данные по нагрузке.
    :return: Индексы начала и конца полки.
    """
    start_shelf = 0  # Индекс элемента начала полки.
    end_shelf = 0  # Индекс элемента конца полки.
    counter = 10  # Счётчик для отхода от индекса начала полки.
    new_start = None  # Индекс самого первого экстремума.
    for i in range(1000, len(burden)):
        # Стартует с 1000-го значения, т.к. до этого мы точно ничего не ждём.
        if all(burden[i] > j for j in burden[i - 10:i]) and all(burden[i] > j for j in burden[i + 1 : i + 10]):
            new_start = i
            break

    for i in range(new_start, len(burden)):  # Цикл для поиска начала и конца полки.
        if not start_shelf:  # Пока не задано значение начала полки.
            if len({el // 100 for el in burden[i:i + 10]}) == 2:
            # Если длина можества из 10-ти следущих значений равна 2.
                start_shelf = i  # Значит, спустя 10 значений будет начало полки.

        else:  # Если мы знаем значение начала полки.
            if counter:  # Пропускаем несколько циклов.
                counter -= 1
                continue
            if len({el // 100 for el in burden[i:i + 10]}) > 3:
            # Если длина множества из 10-ти следующих значений больше 3.
                end_shelf = i  # Значит, найден конец полки.
                break  # Остановка цикла после нахождения конца полки.

    return start_shelf, end_shelf


def search_extremes(elng: list[float], end_shelf: int) -> list[int]:
    """
    Функция для поиска экстемумов.
    :param elng: Список со значениями удлинения.
    :param end_shelf: Индекс конца полки.
    :return: Скисок индексов экстремумов.
    """
    inds_extremes = []  # Список для точек экстремума.
    counter = 0  # Счётчик для отхода от экстремума, на случей если значение экстремума на графике повторяется.

    for i, el in enumerate(elng[end_shelf:-10], start=end_shelf):  # Цикл для нахождения экстремумов.
        if counter:
            counter -= 1
        # Проверка условия, что экстремум найден, даже если рядом точка с идентичными данными.
        elif ((el <= elng[i - 1] and el < elng[i - 2] and el < elng[i - 3] and el < elng[i - 4] and
               el <= elng[i + 1] and el < elng[i + 2] and el < elng[i + 3] and el < elng[i + 4]) or
              (el >= elng[i - 1] and el > elng[i - 2] and el > elng[i - 3] and el > elng[i - 4] and
               el >= elng[i + 1] and el > elng[i + 2] and el > elng[i + 3] and el > elng[i + 4])):
            inds_extremes.append(i)
            counter = 10

    inds_extremes.append(len(elng))  # Добавление к экстремумам индекса последнего значения.

    return inds_extremes


def search_antipods(list_exstremums: list[int], elongation: list[float]) -> list[int]:
    """
    Функция для поиска аналогичного значения предыдущему верхнему экстремуму, на линии нагружения.
    :param list_exstremums: Экстремумы.
    :param elongation: Список со значениями удлинения.
    :return: Список индексов антиподного значения.
    """
    antipods = []  # Контейнер для антиподов на линии нагружения.
    double_list_exstremums = list_exstremums[:]
    for ind, el in enumerate(elongation[list_exstremums[0]: -100],  # Цикл для нахождения антиподных значений.
                             start=list_exstremums[0]):
        if ind < double_list_exstremums[1]:
            pass
        else:
            if round(el, 2) == round(elongation[double_list_exstremums[0]], 2):
                antipods.append(ind)
                double_list_exstremums = double_list_exstremums[2:]

    return antipods


def search_area_under_line(start_ind: int, end_ind: int, elongation: list[float], burden: list[float]) -> int:
    """
    Функция считает площадь под кривой.
    :param start_ind: Индекс начала кривой.
    :param end_ind: Индекс конца кривой.
    :param elongation: Список удлинения.
    :param burden: Список нагрузки.
    :return: Площадь под заданой кривой.
    """
    need_burden = [i - min(burden[start_ind:end_ind]) for i in
                   burden[start_ind:end_ind]]  # Скорректированные занчения нагрузки.
    need_elongation = elongation[start_ind:end_ind]  # Копия значений удлинения.
    area = 0  # Площадь под кривой.
    for i in range(len(need_elongation) - 1):  # Цикл для подсчёта площадь под кривой методом интегрирования.
        area += (need_burden[i] + need_burden[i + 1]) / 2 * (need_elongation[i + 1] - need_elongation[i])

    return area


def dissipation_coefficient_and_rigidity(total_list: list, elongation: list[float], burden: list[float]) -> (
        float, int):
    """
    Функция для расчёта коэффициента рассеивания и жёсткости по петлям гистерезиса, начиная с амплитуды 0,2.
    :param total_list: список контрольных точек (экстремумов и антиподов).
    :param elongation: Список удлинения.
    :param burden: Список нагрузки.
    :return: Коэффициент рассеивания и жёсткость.
    """
    list_dissipation_coefficient = []  # Коэффициент рассеивания.
    list_rigidity = []  # Жёсткость.
    for i in range(0, len(total_list), 3):  # Цикл для вычисления коэффициента рассеивания и жётскости.
        var_1 = abs(search_area_under_line(total_list[i], total_list[i + 1], elongation, burden))  # Площадь разгружения.
        var_2 = search_area_under_line(total_list[i + 1], total_list[i + 2], elongation, burden)  # Площадь нагружения.
        loop_area = var_2 - var_1  # Площадь петли.
        sum_area = var_1 + var_2  # Сумма площадей под линией разгружения и нагружения.
        dissipation_coefficient = 4 * loop_area / (sum_area * 0.5)  # Вычисление коэффициента рассеивания.

        rigidity = ((burden[total_list[i]] - burden[total_list[i + 1]]) /       # Вычисление жёсткости.
                    (elongation[total_list[i]] - elongation[total_list[i + 1]]))
        list_dissipation_coefficient.append(round(dissipation_coefficient, 4))
        list_rigidity.append(int(rigidity))

    return list_dissipation_coefficient, list_rigidity


def amplitude_calculation(elongation: list[float], total: list[int]):
    """
    Функция для расчёта амплитуды петель.
    :param elongation: Удлинение.
    :param total: Список контрольных точек.
    :return: Список амплитуд.
    """
    amplitudes = []
    # Цикл для нахождения амплитуд по петлям путём раззости
    for i in range(0, len(total), 3):
        amplitudes.append(round((elongation[total[i]] - elongation[total[i + 1]]) / 2, 2))
    return amplitudes


def data_counting_for_write(link: str, num_list: int, using=None) -> (list, list):
    """
    Функция возвращающая занчения коэффициента рассеивания и жёсткости в списочном формате отдельно выбраного листа.
    :param link: Ссылка на файл.
    :param num_list: Номер листа.
    :param using: Строка обозначающая, что с данными хочет сделать пользователь,
    просмотреть, либо занести в таблицу ("show" / "download").
    :return: Кортеж из двух списков, а именно коэффициента рассеивания и жёсткости,
    либо печать всех контрольных значений по листу.
    """
    the_elongation, the_burden = read_and_write_xlsx(link, num_list)  # Значения удлинения и нагрузки.
    start_shelf, end_shelf = search_shelf(the_burden)  # Начало и конец полки.
    avg_shelf = (end_shelf + start_shelf) // 2  # Среднее значение для записи в файл.
    if (start_shelf and end_shelf) == 0:  # Проверка на наличие полки.
        return "ErrorShelf"
    extremes = search_extremes(the_elongation, end_shelf + 10)  # Индексы экстремумов.
    antipods = search_antipods(extremes, the_elongation)  # Индексы антиподов.
    total_list = sorted(extremes + antipods)[:-2]  # Список с конечными точками для расчётов.
    # Список амплитуд.
    the_amplitudes = amplitude_calculation(the_elongation, total_list)
    # Расчёт занчений коэф. расс. и жёсткости.
    dissipation_coefficient, rigidity = dissipation_coefficient_and_rigidity(total_list, the_elongation, the_burden)
    #  В зависимости от выбора пользователя, либо печатаем все контрольные значения,
    #  либо возвращаем финальный результат расчётов.
    if using == "show":
        return (f"Начало полки: {start_shelf}\n"  # Печатаем списки найденых значений для расчётов.
            f"Конец полки: {end_shelf}\n"
            f"Экстремумы: {extremes}\n"
            f"Антиподы: {antipods}\n"
            f"Коэффициент рассеивания: {dissipation_coefficient}\n"
            f"Жёсткость: {rigidity}")
    elif using == "write":
        # возвра занчений коэф. расс. и жёсткости.
        return (dissipation_coefficient,
                rigidity,
                the_elongation[avg_shelf],
                the_burden[avg_shelf],
                the_amplitudes,)

def create_xlsx_file_for_write(link_to_file: str, count_loop: int):
    """
    Функция для создания не заполненного результатами файла формата .xlsx.
    :param link_to_file: Ссылка на файл.
    :param count_loop: Количество петель.
    :return: Ссылка на заполняемый файл, количество листов для записи и статус их количества.
    """
    book = openpyxl.Workbook()  # Создаём (пока безымянный) новый документ формата .xlsx для записи результатов.
    count_sheets = input(
        "Напишите количество листов в файле, которые вы бы хотели отправить на вычисления. \nОтвет: ")
        # Принимаем ответ от пользователя.
    if not count_sheets.isdigit():  # Проверка ввода на интовость.
        print(reassurance())  # Печать мотивашки.
        # Повторный вызов функции в случае, если данное число не соответствует предложенным требованиям ввода.
        create_xlsx_file_for_write(link_to_file, count_loop=8)
    count_sheets = int(count_sheets)  # Переопределяем переменную как int.
    check_count_sheets = openpyxl.open(link_to_file, read_only=True)  # Открываем файл с данными на чтение.
    count_sheets = count_sheets if len(check_count_sheets.worksheets) >= count_sheets else len(
        # Если количество листов на запись превышает общее их количество
        check_count_sheets.worksheets)  # в документе, переопределяем переменную на их общее количество.
    status = None
    if count_sheets == len(check_count_sheets.worksheets): status = "all"  # Проверка количества листов для записи.
    elif count_sheets < 1: status = "zero"

    # Активация рабочего листа на запись и дальнейшее его заполнение аннотациями к будующим данным.
    sheet = book.active
    sheet.cell(row=3, column=1).value = "Удлинение по полке"
    sheet.cell(row=4, column=1).value = "Жёсткость по полке"
    sheet.cell(row=7, column=1).value = "Коэффициент"
    sheet.cell(row=9 + count_loop, column=1).value = "Жёсткость"
    sheet.cell(row=11 + count_loop * 2, column=1).value = "Амплитуды"
    # Завершение обозначений в файле.
    return book, count_sheets, sheet, status


def write_date_in_new_file(link_to_file: str):
    """
    Функция для записи вычесляемых данных в новый файл формата xlsx.
    :param link_to_file: Ссылка на файл.
    :param count_loop: Ожидаемое количество петель.
    :return: Создаёт файл с результатами вычислений по конкретному листу.
    """
    count_loop = input("Если вам известно количество петель (амплитуд), то можете ввести нужное значение: ")
    # Предоставление пользователю возможность вписать кол-во петель.
    if count_loop.isdigit() and 0 < int(count_loop) < 100:
        count_loop = int(count_loop)
    else:
        count_loop = 20
    # Создание пустого шаблона xlsx.
    book, count_sheets, sheet, status = create_xlsx_file_for_write(link_to_file, count_loop)
    match status:
        case "zero":
            print("Выбрано отрицательное, либо нулевое значение листов для записи."
                  "\nБудет сознан незаполненный шаблон файла.\n")
        case "all":
            print("Выбрано максимальное количество листов для записи.\n")
    # Начало заполнения листа данными по коэффициенту рассеивания и жётскости.
    for num_list in range(count_sheets):  # Цикл для прохода по листам и записи вычисляемых значений в файл.
        # Результаты коэффициента рассеивания и жётскости петель ([floats], [floats]).
        result = data_counting_for_write(link_to_file, num_list, "write")
        sheet.cell(row=1, column=num_list + 3).value = f"Лист {num_list + 1}"  # Оглавление столбца.
        if result == "ErrorShelf" or len(result[0]) < 2:
            print(f"Лист №{num_list + 1:3d} ❌ ErrorShelf")  # Индикация ошибки при поиске полки.
            sheet.cell(row=2, column=num_list + 3).value = " ❌ "
            continue
        sheet.cell(row=3, column=num_list + 3).value = round(result[2], 2)  # Занесение удлинения по полке.
        sheet.cell(row=4, column=num_list + 3).value = round(result[3], 2)  # Занесение жёсткости по полке.
        for ind_el in range(1, len(result[0]) + 1):  # Цикл для занесения данных по столбцам поэлементно.
            # Занесение результатов коэффициента рассеивания.
            sheet.cell(row=ind_el + 6, column=num_list + 3).value = result[0][ind_el - 1]
            # Занесение результатов жёсткости.
            sheet.cell(row=ind_el + 8 + count_loop, column=num_list + 3).value = result[1][ind_el - 1]
            # # Занесение результатов амплитуды.
            sheet.cell(row=ind_el + 10 + count_loop * 2, column=num_list + 3).value = result[4][ind_el - 1]
        if all(map(lambda x: 0 < x < 3, result[0])):
            print(f"Лист №{num_list + 1:3d} ✔️")  # Индикация успешной записи результатов.
            sheet.cell(row=2, column=num_list + 3).value = " ✔️"  # Занесение пометки в файл.
        else:
            print(f"Лист №{num_list + 1:3d} ❗️ErrorDissipationCoefficient")  # Индикация ошибки коэффициента рассеивания.
            sheet.cell(row=2, column=num_list + 3).value = " ❗️"  # Занесение пометки в файл.
    name_new_file = input("Напишите, желаемое название нового файла: ")  # Ввод желаемого названия файла.
    # Сохранение с условием если пользователь сам проставит расширение.
    book.save(name_new_file + ".xlsx" if name_new_file[-5:] != ".xlsx" else name_new_file)
    book.close()  # Завершение записи и сохранение документа.

    question_3 = input("Желаете продолжить? Да / Нет \nОтвет: ")
    if question_3 not in ("нет", "н", "no", "n"):
        return search_file()  # Повторный вызов функции.

    print("Всего доброго! \nДо новых встреч!")
    exit()  # Выход из программы.

# For_programm/гр1-серия3-001
# гр1-серия3-101


def sheet_data(link_to_file: str):
    """
    Функция для просмотра данных по конкретному листу.
    :param link_to_file: Ссылка на файл.
    :return: Индексы полки, экстремумы, антиподы, коэффициент рассеивания и жёсткость петель по конкретному листу.
    """
    check_count_sheets = openpyxl.open(link_to_file, read_only=True)  # Открываем файл на чтение.
    number_sheet = input("Введите номер листа.\nОтвет: ")  # Предложение ввода номера интерисующего листа.
    # Провека наличия листа под данным номером в файле.
    if number_sheet.isdigit() and 0 < int(number_sheet) <= len(check_count_sheets.worksheets):
        #  Вывод точной информации по конкретному листу на экран.
        print(data_counting_for_write(link_to_file, int(number_sheet) - 1, "show"))
        # Предложение дальнейшего развития событий из четырёх вариантов.
        question = input("\nЕсли хотите посмотреть другой лист из данного файла, нажмите 1.\n"  
                         "Если хотите записать расчёты данных в новый файл, нажмите 2.\n"
                         "Если хотите обработать новый файл, нажмите 3.\n"
                         "Если хотите выйти из программы, нажмите на любую другую клавишу.\nОтвет: ")
        match question:  # Условия на развитие дальнейших событий.
            case "1":
                return sheet_data(link_to_file)  # Повторный вызов функции.
            case "2":
                return write_date_in_new_file(link_to_file)  # Переход на функции по записи данного файла.
            case "3":
                return search_file()  # Переход на этап выбора файла с данными.
            case _:
                print("Всего доброго! \nДо новых встреч!")
                exit()  # Завершение программы.

    else:  # В случае, если пользователь ввёл номер несуществующего листа в документе, функция вызывается вновь.
        print("В данном файле, листа под таким номером не существует.\n")
        print(reassurance())  # Печать мотивашки.

        return sheet_data(link_to_file)  # Повторный вызов функции.


def choice(link_to_file: str):
    """
    Функция для выбора дейтсвия.
    :param link_to_file: Ссылка на файл.
    :return: Зависит от выбора пользователя.
    """
    # Предложение вариантов и ввод от пользователя.
    two_roades = input("Выберите из двух вариантов введя соответствующую цифру (1 / 2):\n"  
                       "1. Записть результатов в новый файл.\n"
                       "2. Демонстрация информации по конкретному листу.\n"
                       "Ответ: ")
    match two_roades:  # Условие на варианты ответа.
        case "1":
            # Переход на функцию по переносу результатов из файла в новый документ.
            return write_date_in_new_file(link_to_file)
        case "2":
            # Переход на функцию для просмотра данных по тому или иному листу.
            return sheet_data(link_to_file)
        case _:
            print(reassurance())  # Печать мотивашки.
            choice(link_to_file)  # Повторный вызов функции.


def search_file():
    """
    Функция для ввода файла для на чтение.
    :return: Следующий шаг работы с файлом.
    """
    try:  # Реализация данной функции через работу с исключениями.
        # Ввод пути к файлу от пользователя.
        link_to_file = input("Напишите название файла до точки или путь к нему.\nОтвет: ")
        # Проверка тернарным оператором наличие / отсутствие расширения в названии.
        link_to_file = link_to_file + ".xlsx" if link_to_file[-5:] != ".xlsx" else link_to_file
        assert openpyxl.open(link_to_file, read_only=True)  # Проверка наличия данного файла в доступной директории.
    except FileNotFoundError:  # Исключение срабатывающее в случае, если файл не найден.
        print("Файл с таким именем не найден.\n")  # Печать причины ошибки.
        print(reassurance())  # Печать мотивашки.
        search_file()  # Повторный вызов функции.
    except Exception:
        # Исключения на всевозможные другие непредвиденные некоректные вводы от пользователя (наример, "Enter").
        print("Почему вы продолжаете бороться мистер андерсон?\n")  # Печать мотивашки.
        search_file()  # Повторный вызов функции.
    else:
        # Если ввод прошёл успешно, пользователь переходит на ступень, выбора работы с файлом.
        return choice(link_to_file)


def start_program() -> search_file:
    """
    Функция приветствия.
    Приветствует пользователя и утонят у него знаком ли он с программой и её функционалом.
    :return: Вызов функции на ввод файла для чтения.
    """
    print(greetings)
    hello_question = input(
        "Знаете ли вы, как пользоваться данной программой? Да / Нет \nОтвет: ")  # Вопрос пользователю.
    print()
    answer = True if hello_question.lower() not in (
        "нет", "н", "no", "n") else False  # Вывод сообщения, если ответ отрицательный.
    if not answer:
        print(familiarization)
    return search_file()


def reassurance():
    """
    Функция возвращающая случайный мотивирующий фразеологизм.
    :return: Случайный фразеологизм.
    """
    from random import choice
    set_of_phrases = (
        "Ошибки - наши лучшие учителя!\n",
        "Хмели сумели и ты сможешь!\n",
        "Всегда можно начать всё заново!\n",
        "Силён не тот кто не падал, а тот кто находил в себе силы поднятся!\n"
    )
    return choice(set_of_phrases)


if "__main__" == __name__:
    start_program()
